from random import choice
from verb_body import *
from database_body import *
import openpyxl
from openpyxl import load_workbook


class MainList:
    DATA = [[('1rst, sg', '-mi', '-i'), ('1rst, du', '-vas', '-vahe'), ('1rst, pl', '-mas', '-mahe')],
            [('2nd, sg', '-si', '-se'), ('2nd, du', '-thas', '-ithe'), ('2nd, pl', '-tha', '-dhve')],
            [('3rd, sg', '-ti', '-te'), ('3rd, du', '-tas', '-ite'), ('3rd, pl', '-nti', '-nte')]]

    RAND_VERB_RES = ''
    RAND_CONJ_RES = ''

    def get_random_conj(self) -> tuple:
        all_data = [elem for item in MainList.DATA for elem in item]
        res = choice(all_data)
        self.RAND_CONJ_RES = res
        return res

    def show_conjugation_hint(self):
        if self.RAND_VERB_RES[-1] == 'P':
            return self.RAND_CONJ_RES[-2]
        elif self.RAND_VERB_RES[-1] == 'A':
            return self.RAND_CONJ_RES[-1]
        elif self.RAND_VERB_RES[-1] == 'U':
            return self.RAND_CONJ_RES[-2:]

    @classmethod
    def validate_data(cls, word):
        return type(word) == str

    @staticmethod
    def check_availability(data):
        flag = True
        wb = openpyxl.load_workbook('sample.xlsx')
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == data:
                        flag = False
        return flag

    def get_random_verb(self):
        with sq.connect('verb_base.db') as con:
            cur = con.cursor()
            res = cur.execute("""SELECT * FROM verb_base""")
            con.commit()

        self.RAND_VERB_RES = choice([*res])
        return self.RAND_VERB_RES

    # def add_data(self, verb: 'Verb'):
    #     with sq.connect('verb_base.db') as con:
    #         cur = con.cursor()
    #         if all(map(lambda x: bool(x), [verb.inf_value, verb.pres_base, verb.translation, verb.conj])):
    #             cur.execute(f"""INSERT INTO verb_base VALUES (
    #             '{verb.inf_value}', '{verb.pres_base}', '{verb.translation}', '{verb.conj}')""")
    #         con.commit()


    # def add_data(self, ent1,ent2,ent3,ent4):
    #     with sq.connect('verb_base.db') as con:
    #         cur = con.cursor()
    #         if all(map(lambda x: bool(x), [ent1,ent2,ent3,ent4])):
    #             cur.execute(f"""INSERT INTO verb_base VALUES (
    #             '{ent1}', '{ent2}', '{ent3}', '{ent4}')""")
    #         con.commit()

# if __name__ == '__main__':
#     a = MainList()
#     a1 = Verb('-/is./-', '-/iccha-/-', '-/хотіти, воліти/-', '-/U/-')
#     a2 = Verb('-/likh/-', '-/likha/-', '-/писати/-', '-/P/-')
#     a3 = Verb('-/khaad/-', '-/khaada/-', '-/їсти/-', '-/P/-')
#     a4 = Verb('-/iiks./-', '-/iiks.a/-', '-/бачити, дивитись/-', '-/A/-')
#     a5 = Verb('-/c,iks./-', '-/c,iks.a/-', '-/навчатись/-', '-/A/-')
#     a6 = Verb('-/khaad/-', '-/khaada/-', '-/їсти/-', '-/P/-')
#     a(a1)
#     a(a2)
#     a(a3)
#     a(a4)
#     a(a5)
#     a(a6)
# print('a.generate_random_data', a.generate_random_data())
# # print('a.get_random_verb():', a.get_random_verb())
# print('a.show_random_verb():', a.show_random_verb())
# # print('a.get_random_data()', a.get_random_data())
# print('a.show_random_data()', a.show_random_conj())
# print('a.RAND_VERB_RES:', a.RAND_VERB_RES)
# print('a.RAND_CONJ_RES:', a.RAND_CONJ_RES)
# print('a.show_conjugation_hint', a.show_conjugation_hint())
# print('a.show_verb_hint', a.show_verb_hint())
